#!/usr/bin/env python3
"""Manpower-loaded schedule generator (xlsx).

Turns a list of activities (trade, start week, duration, peak crew, load curve)
into a weekly manpower-loaded schedule workbook:

  - Activities sheet: per-activity crew loaded across each week, with a left
    color spine by trade, plus a TOTAL manpower row and man-hours summary.
  - A manpower histogram chart (the classic crew-curve) driven by the totals.

Load curves supported per activity:
  flat      constant peak crew for the duration
  ramp      linear ramp up then down (triangular), peaking mid-span
  front     front-loaded (peak early, taper off)
  back      back-loaded (ramp to peak at the end)

Usage:
  python manpower_schedule.py                         # writes the built-in sample
  python manpower_schedule.py --in plan.json --out schedule.xlsx
  python manpower_schedule.py --hours-per-week 40

`plan.json` shape:
  {
    "project": "Project Alpha — Level 3 Buildout",
    "prepared_by": "Gavin Watson",
    "weeks": 10,
    "hours_per_week": 40,
    "activities": [
      {"name":"Demo & layout","trade":"Laborers","start":1,"duration":2,"peak":6,"curve":"front"},
      ...
    ]
  }
"""
from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Trade -> spine color (ARGB, no leading #).
TRADE_COLORS = {
    "Laborers": "FF94A3B8",
    "Pipefitters": "FF3B82F6",
    "Plumbers": "FF22C55E",
    "Sheet Metal": "FFF59E0B",
    "Electricians": "FFEAB308",
    "Controls": "FFA855F7",
    "Insulators": "FF14B8A6",
    "TAB": "FFEF4444",
}
DEFAULT_COLOR = "FF64748B"

# A reasonable demo plan (mirrors the "Level 3 buildout" task in the app).
SAMPLE = {
    "project": "Project Alpha — Level 3 Buildout (Mechanical)",
    "prepared_by": "Gavin Watson",
    "weeks": 10,
    "hours_per_week": 40,
    "activities": [
        {"name": "Demo & layout", "trade": "Laborers", "start": 1, "duration": 2, "peak": 6, "curve": "front"},
        {"name": "Hangers & supports", "trade": "Pipefitters", "start": 2, "duration": 3, "peak": 8, "curve": "ramp"},
        {"name": "Chilled water piping", "trade": "Pipefitters", "start": 3, "duration": 5, "peak": 14, "curve": "ramp"},
        {"name": "Plumbing rough-in", "trade": "Plumbers", "start": 3, "duration": 4, "peak": 6, "curve": "flat"},
        {"name": "Ductwork mains", "trade": "Sheet Metal", "start": 4, "duration": 4, "peak": 10, "curve": "ramp"},
        {"name": "Branch ducts & devices", "trade": "Sheet Metal", "start": 6, "duration": 3, "peak": 6, "curve": "back"},
        {"name": "Equipment set", "trade": "Pipefitters", "start": 5, "duration": 2, "peak": 5, "curve": "flat"},
        {"name": "Controls rough-in", "trade": "Controls", "start": 6, "duration": 4, "peak": 4, "curve": "ramp"},
        {"name": "Insulation", "trade": "Insulators", "start": 7, "duration": 3, "peak": 5, "curve": "back"},
        {"name": "Test & balance", "trade": "TAB", "start": 9, "duration": 2, "peak": 3, "curve": "back"},
    ],
}


def crew_curve(peak: int, duration: int, curve: str) -> list[int]:
    """Return a per-week crew list of length `duration` following the curve shape."""
    if duration <= 0:
        return []
    if duration == 1:
        return [peak]
    out = []
    for i in range(duration):
        frac = i / (duration - 1)  # 0..1 across the span
        if curve == "flat":
            v = peak
        elif curve == "front":
            v = peak * (1 - 0.6 * frac)
        elif curve == "back":
            v = peak * (0.4 + 0.6 * frac)
        else:  # ramp (triangular, peak in the middle)
            v = peak * (1 - abs(2 * frac - 1) * 0.5)
        out.append(max(1, int(round(v))))
    # ensure the peak is actually hit somewhere
    if out and max(out) < peak:
        out[out.index(max(out))] = peak
    return out


def build(plan: dict, out_path: Path) -> dict:
    weeks = int(plan.get("weeks", 10))
    hpw = int(plan.get("hours_per_week", 40))
    activities = plan["activities"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Manpower Schedule"

    thin = Side(style="thin", color="FF333B49")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="FF10141D")
    hdr_font = Font(bold=True, color="FFE7EBF2", size=10)
    body_font = Font(color="FFE7EBF2", size=10)
    mut_font = Font(color="FF8A94A6", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # --- Title block ---
    ws["A1"] = plan.get("project", "Manpower-Loaded Schedule")
    ws["A1"].font = Font(bold=True, size=14, color="FF34D399")
    ws["A2"] = f"Prepared by {plan.get('prepared_by','')}   ·   {weeks} weeks   ·   {hpw} hrs/week"
    ws["A2"].font = mut_font

    # --- Header row ---
    head_row = 4
    fixed = ["Activity", "Trade", "Start", "Dur", "Peak"]
    for c, label in enumerate(fixed, start=1):
        cell = ws.cell(row=head_row, column=c, value=label)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center
    for w in range(1, weeks + 1):
        cell = ws.cell(row=head_row, column=5 + w, value=f"Wk {w}")
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center
    mh_col = 5 + weeks + 1
    cell = ws.cell(row=head_row, column=mh_col, value="Man-hrs")
    cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center

    # --- Activity rows ---
    first_data = head_row + 1
    week_totals = [0] * (weeks + 1)  # 1-indexed
    for r_off, act in enumerate(activities):
        r = first_data + r_off
        trade = act.get("trade", "")
        color = TRADE_COLORS.get(trade, DEFAULT_COLOR)
        start = int(act["start"]); dur = int(act["duration"]); peak = int(act["peak"])
        curve = act.get("curve", "ramp")
        loads = crew_curve(peak, dur, curve)

        ws.cell(row=r, column=1, value=act["name"]).font = body_font
        ws.cell(row=r, column=1).alignment = left
        # trade cell carries the color spine
        tcell = ws.cell(row=r, column=2, value=trade)
        tcell.font = body_font; tcell.alignment = left
        tcell.fill = PatternFill("solid", fgColor=color)
        tcell.font = Font(bold=True, color="FF070910", size=9)
        ws.cell(row=r, column=3, value=start).alignment = center
        ws.cell(row=r, column=4, value=dur).alignment = center
        ws.cell(row=r, column=5, value=peak).alignment = center

        act_mh = 0
        for i, crew in enumerate(loads):
            wk = start + i
            if wk < 1 or wk > weeks:
                continue
            col = 5 + wk
            cell = ws.cell(row=r, column=col, value=crew)
            cell.alignment = center; cell.font = body_font; cell.border = border
            # shade by intensity of the trade color
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(color="FF070910", size=9, bold=crew == peak)
            week_totals[wk] += crew
            act_mh += crew * hpw
        ws.cell(row=r, column=mh_col, value=act_mh).alignment = center
        ws.cell(row=r, column=mh_col).font = body_font

        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border

    # --- TOTAL manpower row ---
    total_row = first_data + len(activities)
    tcell = ws.cell(row=total_row, column=1, value="TOTAL MANPOWER")
    tcell.font = Font(bold=True, color="FF34D399", size=10)
    ws.cell(row=total_row, column=5, value=max(week_totals)).font = Font(bold=True, color="FF34D399")
    grand_mh = 0
    for wk in range(1, weeks + 1):
        col = 5 + wk
        cell = ws.cell(row=total_row, column=col, value=week_totals[wk])
        cell.alignment = center
        cell.font = Font(bold=True, color="FFE7EBF2", size=10)
        cell.fill = hdr_fill; cell.border = border
        grand_mh += week_totals[wk] * hpw
    ws.cell(row=total_row, column=mh_col, value=grand_mh).font = Font(bold=True, color="FF34D399")

    # --- Summary block ---
    s = total_row + 2
    summary = [
        ("Peak manpower", f"{max(week_totals)} workers (Wk {week_totals.index(max(week_totals))})"),
        ("Total man-hours", f"{grand_mh:,}"),
        ("Avg crew / active week", f"{round(sum(week_totals)/max(1,sum(1 for t in week_totals if t)),1)}"),
        ("Duration", f"{weeks} weeks"),
    ]
    for i, (k, v) in enumerate(summary):
        ws.cell(row=s + i, column=1, value=k).font = mut_font
        ws.cell(row=s + i, column=3, value=v).font = body_font

    # --- Column widths ---
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    for c in range(3, 6):
        ws.column_dimensions[get_column_letter(c)].width = 6
    for w in range(1, weeks + 1):
        ws.column_dimensions[get_column_letter(5 + w)].width = 6
    ws.column_dimensions[get_column_letter(mh_col)].width = 10
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(row=first_data, column=6)

    # --- Manpower histogram (crew curve) on a second sheet ---
    chart_ws = wb.create_sheet("Crew Curve")
    chart_ws["A1"] = "Week"; chart_ws["B1"] = "Total Manpower"
    for wk in range(1, weeks + 1):
        chart_ws.cell(row=1 + wk, column=1, value=f"Wk {wk}")
        chart_ws.cell(row=1 + wk, column=2, value=week_totals[wk])
    chart = BarChart()
    chart.type = "col"; chart.title = "Manpower Loading Curve"
    chart.y_axis.title = "Workers"; chart.x_axis.title = "Week"
    chart.height = 8; chart.width = 22
    data = Reference(chart_ws, min_col=2, min_row=1, max_row=1 + weeks)
    cats = Reference(chart_ws, min_col=1, min_row=2, max_row=1 + weeks)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    chart_ws.add_chart(chart, "D2")
    chart_ws.sheet_view.showGridLines = False

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return {
        "out": str(out_path),
        "peak": max(week_totals),
        "peak_week": week_totals.index(max(week_totals)),
        "total_man_hours": grand_mh,
        "weeks": weeks,
        "activities": len(activities),
    }


def main() -> None:
    ap = argparse.ArgumentParser(description="Generate a manpower-loaded schedule workbook.")
    ap.add_argument("--in", dest="inp", help="plan JSON (defaults to built-in sample)")
    ap.add_argument("--out", dest="out", default=str(Path(__file__).parent / "output" / "manpower_schedule.xlsx"))
    ap.add_argument("--hours-per-week", type=int, default=None)
    args = ap.parse_args()

    plan = json.loads(Path(args.inp).read_text()) if args.inp else dict(SAMPLE)
    if args.hours_per_week:
        plan["hours_per_week"] = args.hours_per_week

    res = build(plan, Path(args.out))
    print(json.dumps(res, indent=2))


if __name__ == "__main__":
    main()
